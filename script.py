# -*- coding: utf-8 -*-
"""
Created on Sat Nov 14 14:14:53 2020

@author: Matthew
"""

import os
import numpy as np
import cv2
import pytesseract as tess
tess.pytesseract.tesseract_cmd = r'C:\Users\Matthew\AppData\Local\Tesseract-OCR\tesseract.exe'
from openpyxl import Workbook
import math
import glob

#Calculate the distance between a designated box to every other box, returns the array containing the distances
def calculateDistance(coord_arr,a):
    dist_arr = []
    midpoint_arr = []
    
    for x in range(0,(len(coord_arr)),1):
        midpoint = (int((coord_arr[x][0][0] + coord_arr[x][3][0])/2),int((coord_arr[x][0][1] + coord_arr[x][3][1])/2))
        midpoint_arr.append(midpoint)
    
    for y in range(0,len(midpoint_arr),1):      
        distance = math.sqrt((((midpoint_arr[a][0] - midpoint_arr[y][0]) ** 2) + ((midpoint_arr[a][1] - midpoint_arr[y][1]) ** 2)))
        distance = int(distance)
        
        if distance == 0:
            distance += 1111
                    
        dist_arr.append(distance)
    
    return dist_arr

#Calculates the midpoint of each box, returns the array containing all the midpoint of each box    
def calculateMidpoint(coord_arr):
    midpoint_arr = []
    
    for x in range(0,(len(coord_arr)),1):
        midpoint = (int((coord_arr[x][0][0] + coord_arr[x][3][0])/2),int((coord_arr[x][0][1] + coord_arr[x][3][1])/2))
        midpoint_arr.append(midpoint)
        
    return midpoint_arr

#Calculate the number of rows in the engineering drawing with reference to the first field title / heading
def getRow(bU_arr):
  rows = []
  current_row = []
  y = int(bU_arr[0][7]) + (int(bU_arr[0][9])//2) #Mid-y
  for b_idx, b in enumerate(bU_arr):
    current_y = int(b[7]) + (int(b[9])//2)
    if np.abs(y - current_y) < 30:
      current_row.append(b)
    else:
      rows.append(current_row)
      current_row = []
      current_row.append(b)
      y = int(b[7]) + (int(b[9])//2)
  rows.append(current_row)
  return rows

#Perform image filtering, prepares the engineering drawing for text scanning
def processImg(img):
  img_bi = cv2.threshold(img,0,255,cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
  mask = ~img_bi
  # cv2.imwrite("step0-img_bi.jpg", img_bi) 
  
  [nrow,ncol] = img_bi.shape
  
  kernel_len = np.array(img).shape[1]//100
  hor_SE = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len,1))
  vert_SE = cv2.getStructuringElement(cv2.MORPH_RECT, (1,kernel_len))
  
  process_img = cv2.morphologyEx(mask, cv2.MORPH_OPEN, sE_diamond1)
  
  process_img = cv2.morphologyEx(process_img, cv2.MORPH_OPEN, sE_square1, iterations=3)
  #cv2.imwrite("step1_process_img.jpg", process_img)
  step1 = cv2.medianBlur(process_img,13)
  #cv2.imwrite("step1.jpg", step1) 
  
  # Remove figure
  step2 = cv2.bitwise_xor(step1,img_bi)
  step2 = cv2.erode(~step2, sE_square1, iterations=3)
  step2 = cv2.dilate(step2, sE_square3, iterations=4)
  step2 = cv2.dilate(step2, sE_square2, iterations=12)
  
  drawings = cv2.bitwise_and(~step2, img) + step2
  
  step3 = cv2.bitwise_and(img_bi, step2)
  #cv2.imwrite("step4.jpg", step4)
  
  # Get lines from table
  process_img = cv2.erode(~step3, hor_SE, iterations=3)
  hor_line = cv2.dilate(process_img, hor_SE, iterations=4)
  
  process_img = cv2.erode(~step3, vert_SE, iterations=3)
  vert_line = cv2.dilate(process_img,vert_SE, iterations=4)

  step4 = hor_line + vert_line
  
  step5 = cv2.bitwise_or(step3,step4)
  #cv2.imwrite("step6.jpg", step6)
  
  return step5, drawings

#The algorithm to extract and save the correct data into their corresponding fields in the matching array
def getData(boxes):
  b_arr = []
  coord_arr = []
  
  #Scans the data using pyTesseract.
  for b_idx,b in enumerate(boxes.splitlines()):
      if b_idx!=0:
          b = b.split()
      if len(b)==12:
          x,y,w,h = int(b[6]),int(b[7]),int(b[8]),int(b[9])
          cv2.rectangle(img2,(x,y),(w+x,h+y),(0,0,255),2)
          if len(b[11]) >= 1 and b[11] != '|':
              b_arr.append(b) #Stores the data of each box scanned into an array. 
              coord_arr.append(((x,y),(x,(y+w)),((x+h),y),((x+h),(y+w)))) #Stores the coordinate of each box scanned into an array
              #cv2.putText(img2,str(len(b_arr)),(x,y),cv2.FONT_HERSHEY_COMPLEX,1,(40,40,100),2)
              #cv2.putText(img2,b[11],(x,y),cv2.FONT_HERSHEY_COMPLEX,1,(40,40,100),2)
  #cv2.imwrite("step7.jpg", img2)   
  
  #Initialising arrays for the process
  #the index arrays contains the index of boxes in b_arr
  index_arr = []
  index_arr_2 = []
  index_arr_2_final = []
  
  #temporary arrays contain the index of boxes that do not belong in the amendments table
  temp = []
  temp_2 = []
  
  #the box Updated arrays. bU_arr will contain all of the data excluding the amendments table data
  #bU_amend_arr contains all of the data that belongs in the amendments table
  bU_arr = []
  bU_amend_arr = []
  bUH_arr = []
  
  #Fixing some outlier issue, helps with improving the algorithm just a bit
  for z in range(0, len(b_arr),1):
      for x in ["|PROJECT"]:
          if any(x in s for s in b_arr[z]) == True:
              b_arr[z][11] = "PROJECT"
  
  for z in range(0, len(b_arr),1):
      for x in ["|COMPANY"]:
          if any(x in s for s in b_arr[z]) == True:
              b_arr[z][11] = "COMPANY"
              
  for z in range(0, len(b_arr),1):
      for x in ["APPROVFD"]:
          if any(x in s for s in b_arr[z]) == True:
              b_arr[z][11] = "APPROVED"            
  
  #Cancelling out the second box. For example: Drawing and Number:, if they are next to each other, Number: becomes " "
  for i in range(0,len(b_arr),1):
      for x in ["DRAWING","DRAWN","CHECKED","APPROVED","PROJECT","CAD","COMPANY"]:
          if any(x in s for s in b_arr[i]) == True:
              for y in ["NUMBER:","BY:","NO:","NAME:","NO.:"]:
                  if any(y in s for s in b_arr[i]) == True:
                      b_arr[i][11] = (str(x) + " " + str(y)) 
                      
              if " " not in b_arr[i][11]:
                  if "CAD NO:" not in b_arr[i][11]:
                      index_arr.append(i)
                      
                      if b_arr[i] == b_arr[len(b_arr) - 1]:
                          i -= 0
  
  #Store all of the joined field title / headings into a separate array
  bUH_arr = b_arr                      
  
  #Joining the field titles together. For example, Drawing and Number: joined into Drawing Number:
  for i in range(0, len(index_arr),1):
      if b_arr[int(index_arr[i])][11] in ["DRAWING","DRAWN","CHECKED","APPROVED","PROJECT","CAD","COMPANY"]:
          try:
            if b_arr[int(index_arr[i])+1][11] in ["NUMBER:","DRAWN:","TITLE:","BY:","BY:","NO:","NAME:","NO.:"]:
                if str(b_arr[int(index_arr[i])][11]) + " " + str(b_arr[int(index_arr[i])+1][11]) in ["DRAWING NUMBER:","DRAWING TITLE:","DRAWN BY:","CHECKED BY:","APPROVED BY:","PROJECT NO:","CAD NO:","COMPANY NAME:","DRAWING NO.:","DRAWING NO:"]:
                    bUH_arr[int(index_arr[i])][11] = (str(b_arr[int(index_arr[i])][11]) + " " + str(b_arr[int(index_arr[i])+1][11]))
                    bUH_arr[int(index_arr[i])+1][11] = " "
          except:
            continue
  
  #Using the "AMENDMENTS" box as reference, drawing a radius around it. Any box that has a distance lesser than a set radius
  #is considered to be in the amendments table
  for i in range(0,len(b_arr),1):
      if b_arr[i][11] in ["AMENDMENTS"]:
          completeDist = calculateDistance(coord_arr,i)        
          midpoint_arr = calculateMidpoint(coord_arr)
          if (midpoint_arr[i][0] < (int((ncol)*0.2))) or (int((ncol)*0.4)) < midpoint_arr[i][0] < (int((ncol)*0.6)) or ((int((ncol)*0.8)) < midpoint_arr[i][0] < (int(ncol)))  or ((int((ncol)*0.6)) < midpoint_arr[i][0] < (int((ncol)*0.8))):
              radius = 450
          elif ((int((ncol)*0.2)) < midpoint_arr[i][0] < (int((ncol)*0.4))):    
              radius = 1000
          
          #The index of every box that supposedly belongs to the amendment tables is stored in the index array, index_arr_2  
          for k in range(0,len(completeDist),1):
              if (completeDist[k] < radius):
                  index_arr_2.append(k)
  
  #Managing offset failure
  # If the amendments table extraction used a radius larger than needed, the unneeded values are appended in temp
  for y in range(len(index_arr_2)):
      if b_arr[int(index_arr_2[y])][11] in ["TITLE:","DRAWING TITLE:"]:
          temp.append(y)
          
  #if temp contains items, it means that the amendment table extraction used a radius larger than needed
  #if temp, then append the new items into temp_2 and append the values that are not in b_arr[int(index_arr_2)[i]]  
  if temp:
      for i in range(0,int(temp[0]),1):
          temp_2.append(index_arr_2[i]) 
          
  if not temp:
      temp_2 = [] 
       
  #if temp_2 is not empty, that means there are values that do not belong to amendments table
  #append only the values that only belongs to amendments table 
  if temp_2:
      for i in range(0, len(temp_2),1):
          index_arr_2_final.append(int(index_arr_2[i]))    
  
      for i in range(0, (len(index_arr_2_final)),1):
          bU_amend_arr.append(b_arr[int(index_arr_2_final[i])])
  
  if not temp_2:
      for i in range(0, (len(index_arr_2)),1):
          bU_amend_arr.append(b_arr[int(index_arr_2[i])])
      
  
  #update the arrays by appending all values that do not belong in the amendments table into bU_arr (box Updated array)  
  if temp_2:
      for idx in range(len(b_arr)):
          if idx not in index_arr_2_final:
              if bUH_arr[idx][11] not in ["AMENDMENTS"," "]:
                  if int(b_arr[idx][9]) > 20 and b_arr[idx][11] not in ['=']:
                      bU_arr.append(b_arr[idx])  
  
  if not temp_2:
      for idx in range(len(b_arr)):
          if idx not in index_arr_2:
              if bUH_arr[idx][11] not in ["AMENDMENTS"," "]:
                  if int(b_arr[idx][9]) > 20 and b_arr[idx][11] not in ['=']:
                      bU_arr.append(b_arr[idx])            

  #get the number of rows of the box array starting from the first field title encountered in the engineering drawing
  rows = getRow(bU_arr)
  all_data = []
  
  header = None
  for r_idx, row in enumerate(rows):
    for box in row:
      if box[11] in headers:
        header = r_idx
        break
    if header != None:
      break
  rows = rows[header:]
  
  #For each row that only contains the field titles, check the next row
  #If the midpoint of the data box is within the boundaries of: 
  #the first field title's left corner and the next field title's right corner - then the data belongs to the first field title
  for row_idx in range(0,len(rows), 2):
    for hb_idx, header_box in enumerate(rows[row_idx]):
      header = header_box[11]
      data = ''
      try:
        for data_box in rows[row_idx+1]:
          data_x = int(data_box[6]) + (int(data_box[8])//2)
          if data_x > int(header_box[6]):
            if hb_idx < len(rows[row_idx])-1: #Check if header not last
              if data_x < int(rows[row_idx][hb_idx+1][6]):
                data += ' ' + data_box[11]
            else: #Last header of row
              data += ' ' + data_box[11]
      except:
        continue
      all_data.append([header, data])
      
  for x in range(0,len(b_arr),1):
    if b_arr[x][11] in ["AMENDMENTS"]:
      all_data.append([b_arr[x][11],"DATA"])

  #Process is repeated for the amendment array
  
  #get the number of rows of the box array starting from the first field title encountered in the engineering drawing
  #ONLY FOR AMENDMENTS TABLE!
  rows = getRow(bU_amend_arr) 
  header = None
  for r_idx, row in enumerate(rows):
    for box in row:
      if box[11] in headers_amend:
        header = r_idx
        break
    if header != None:
      break
  rows = rows[header:]

  #THIS IS FOR 3x3 AMENDMENT TABLE, 3x5 AMENDMENT TABLE FAILS FROM THIS
  for row_idx in range(1,len(rows), 1):
    for hb_idx, header_box in enumerate(rows[row_idx]):     

      data = ''
      
      try:
        header = rows[0][hb_idx][11]
        for data_box in rows[row_idx]:
          
          data_x = int(data_box[6]) + (int(data_box[8])//2)
          if data_x > int(header_box[6]):
            if hb_idx < len(rows[row_idx])-1: #Check if header not last
              if data_x < int(rows[row_idx][hb_idx+1][6]):
                data += ' ' + data_box[11]        
            else: #Last header of row
              data += ' ' + data_box[11]
      except:
        continue
      all_data.append([header, data])

  return all_data 
 
#Structuring Element generation for filtering the image    
sE_square1 = cv2.getStructuringElement(cv2.MORPH_RECT,(3,3))
sE_square2 = cv2.getStructuringElement(cv2.MORPH_RECT,(20,20))
sE_square3 = cv2.getStructuringElement(cv2.MORPH_RECT,(50,50))
sE_diamond1 = np.array([
  [0,0,0,0,1,0,0,0,0],
  [0,0,0,1,1,1,0,0,0],
  [0,0,1,1,1,1,1,0,0],
  [0,1,1,1,1,1,1,1,0],
  [1,1,1,1,1,1,1,1,1],
  [0,1,1,1,1,1,1,1,0],
  [0,0,1,1,1,1,1,0,0],
  [0,0,0,1,1,1,0,0,0],
  [0,0,0,0,1,0,0,0,0]], dtype=np.uint8)

#Setting all the possible field title / headers into an array
headers = ["TITLE:",
           "DRAWING NO:",
           "DRAWING NUMBER:",
           "CONTRACTOR:",
           "DRAWN BY:",
           "CHECKED BY:",
           "APPROVED BY:",
           "CAD NO:",
           "LANG:",
           "PROJECT NO:",
           "PAGE:",
           "STATUS:",
           "STS:",
           "UNIT:",
           "DRAWING TITLE:",
           "DRAWN:",
           "CHECKED:",
           "APPROVED:",
           "COMPANY NAME:",
           "COMPANY:"]

#Setting all the possible field title / headers for amendments table only into an array
headers_amend = ["ISSUE",
                 "DATE",
                 "CHANGE(S)",
                 "BY",
                 "CDK",
                 "REV",
                 "DATE",
                 "BY"]

#Main execution of program
if __name__ == '__main__':
  # Initialise directories
  drawings_path = os.path.join("drawing")
  if not os.path.exists(drawings_path):
      os.makedirs(drawings_path)
      
  workbook_path = os.path.join("workbook")
  if not os.path.exists(workbook_path):
      os.makedirs(workbook_path)
    
  img_path = os.path.join("img")
  _, _, images = next(os.walk(img_path))
  #For specific selection of images to process
  #images = [image[4:] for image in glob.glob(img_path+'/08.png')]
  for image in images:
    img = cv2.imread(os.path.join(img_path, image), 0)
    [nrow,ncol] = img.shape
    
    img2, drawings = processImg(img)
    
    cv2.imwrite(os.path.join(drawings_path, "drawing"+image), drawings)
    
    conf = r'--oem 3 --psm 6'
    boxes = tess.image_to_data(img2,config=conf)
    all_data = getData(boxes)
    #Prints the return value of getData(boxes) into the console
    for d in all_data:
      print(d)
      
    #Initialise a workbook / excel file
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Drawing Info Table"
    
    #Write in to Spreadsheet
    sheet.cell(row=1,column=1).value= "Field Title"
    sheet.cell(row=1,column=2).value= "Content"
    
    for cellrow in range(1,len(all_data)+1):
        sheet.cell(row=cellrow+1,column=1).value = all_data[cellrow-1][0]
        sheet.cell(row=cellrow+1,column=2).value = all_data[cellrow-1][1]
    
    #set the width 25 for A and B column
    for i in range(1,sheet.max_column+1):
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
    
    workbook.save(os.path.join(workbook_path, "excelsheet"+image[:-4]+".xlsx"))