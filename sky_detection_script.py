# -*- coding: utf-8 -*-
"""
Created on Thu Jul  8 04:51:55 2021

@author: Matthew
"""

import cv2
import numpy as np
import os
from openpyxl import Workbook

#4232, 9483, 10917  
#Initialise the pathway to the images
#img_path = os.path.join("images/4232")  
#img_path = os.path.join("images/9483")  
img_path = os.path.join("images/10917")  
_, _, images = next(os.walk(img_path))

#Initialise the pathway to the results folder
#result_path = os.path.join("images/4232-result")
#result_path = os.path.join("images/9483-result")
result_path = os.path.join("images/10917-result")

#Read the mask
ground_truth_path = os.path.join("images/ground_truth")
#ground_truth = cv2.imread(os.path.join(ground_truth_path, "4232.png"), 0)
#ground_truth = cv2.imread(os.path.join(ground_truth_path, "9483.png"), 0)
ground_truth = cv2.imread(os.path.join(ground_truth_path, "10917.png"), 0)

#Make a workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Accuracy & Error rate table"

#Write the headings for each criteria
sheet.cell(row = 1, column = 1).value = "Index"
sheet.cell(row = 1, column = 2).value = "Accuracy"
sheet.cell(row = 1, column = 3).value = "Error rate"
sheet.cell(row = 1, column = 5).value = "Average Accuracy"
sheet.cell(row = 1, column = 6).value = "Average Error rate"

#Used for labelling results
counter = 1

#Main processing algorithm
for image in images:
  img = cv2.imread(os.path.join(img_path, image),1)

  #Calculating the mean intensity value to be used in the Canny edge detection system
  meanIntensityValue = np.mean(img)
  canny = cv2.Canny(img, (0.5 * meanIntensityValue), (1.0 * meanIntensityValue))
  
  #Taking the region that IS NOT the edges. 
  #The edges represent the ground, and the area above the edges are the sky
  sky_mask = (canny < 50).astype(np.uint8)
  
  #Intialise the kernel to remove the noise present in the image
  kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (19, 19))
  #Using the kernel, performing opening to remove the points
  filtered_mask = cv2.morphologyEx(sky_mask, cv2.MORPH_OPEN, kernel)
  
  #Get the size of the image (or from the mask as they are the same)
  [height, width] = filtered_mask.shape
  
  result_mask = np.zeros((height, width), dtype = 'uint8')
  
  #Boundary detection + noise exclusion system
  #Mapping the pixels that belong to the sky based on the first tangible edge the algorithm sees
  for i in range(width):
    raw = filtered_mask[:, i]
    try:
      zero_index_list = []
      #Set the error margin
      #Error margin is to account for possible noise that still exists in the sky
      error_margin = 10
      for x in range(error_margin):
        zero_index_list.append(np.where(raw == 0)[0][x])
        
      #Initially set the index. This assumes that the first one is the true sky.
      #This assumes that the first zero is the last pixel of the sky or the first pixel of the ground
      #It is always assumed that the sky is on top of the image
      first_zero_index = np.where(raw == 0)[0][0]

      for x in range(error_margin - 1):
        #High confidence if the neighbouring pixels are only 1 or 2 apart, it is indeed the ground
        if(zero_index_list[x+1] - zero_index_list[x] > 3):
          #print(i, "Das super wrong")
          #first_zero_index = np.where(raw == 0)[0][x+1]
          first_zero_index = zero_index_list[x+1]
      
      result_mask[:first_zero_index, i] = 1
    except:
      continue
    
  #Accuracy and error rate calculation
  acc_pix = 0
  error_pix = 0
  for x in range(height):  
    for y in range(width):
      if result_mask[x, y] == 1:
        if ground_truth[x,y] > 0:
          acc_pix += 1
        else:
          error_pix += 1
  
  #If condition to account when a fake sky is detected
  if (np.count_nonzero(ground_truth)) != 0:
    accuracy = (acc_pix / np.count_nonzero(ground_truth))
    error_rate = (error_pix / np.count_nonzero(result_mask))
  #This implies that based on the ground truth, there should not be a sky  
  #The accuracy of no sky images should always no 0. It is impossible to incorrectly identify sky pixels if they do not exist
  #For error rate of no sky images, the number of false positive pixels will be calculated
  else:
    accuracy = 0
    error_rate = (error_pix / (len(ground_truth) * len(ground_truth[0])))
  
  #Writing results into table
  sheet.cell(row = 1+counter, column = 1).value = counter
  sheet.cell(row = 1+counter, column = 2).value = accuracy
  sheet.cell(row = 1+counter, column = 3).value = error_rate
  sheet.cell(row = 1+1, column = 5).value = '= AVERAGE(B:B)'
  sheet.cell(row = 1+1, column = 6).value = '= AVERAGE(C:C)'
  
  #Produce the final image (Extracted sky region)
  result = cv2.bitwise_and(img, img, mask=result_mask)
  
  #Saving the results
  workbook.save(os.path.join(result_path, "Accuracy_ErrorRate" + ".xlsx"))
  cv2.imwrite(os.path.join(result_path, "image " + str(counter-1) +".jpg"), result)

  #Incrementing the counter so that the data is written-
  #-into the correct cells inside the excel spreadsheet
  counter += 1